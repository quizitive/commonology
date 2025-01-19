# Alex created a spreadsheet of all the people who made referrals in past games.
# This script reads them form an excel spreadsheet and saves them to
# the commonology database.

import sys
import os
import django
import openpyxl

path = os.getcwd()
sys.path.append(path)
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "project.settings")
django.setup()


from users.models import Player


players = Player.objects.all()

fixes = {
    "Mr. Nobody": "",
    "Mr. Someone": "someone@somebody.edu",
}

fixes = {k.lower(): v for k, v in fixes.items()}


def find_by_name(name):
    name = name.lower().strip()

    if name in fixes:
        email = fixes[name].strip()
        if not email:
            return None
        p = Player.objects.get(email=email)
        return p

    words = name.split()

    last_name = words[-1]
    first_name = words[0]

    for u in players:
        if name == u.display_name.lower():
            return u
        if last_name == u.last_name.lower() and first_name == u.first_name.lower():
            return u
        if name == u.email.lower():
            return u

    return None


def process_column(w, c):
    r = 1
    referrer = w.active.cell(row=r, column=c).value.strip()
    referrer_obj = find_by_name(referrer)

    if not referrer_obj:
        print("Cannot resolve ", referrer)
        return

    while referrer:
        r += 1
        referee = w.active.cell(row=r, column=c).value
        if not referee:
            break

        p = find_by_name(referee)
        if p:
            if p.referrer:
                print(referee, "already referred by", p.referrer, "not changing it to", referrer_obj)
            else:
                p.referrer = referrer_obj
                p.save()
        else:
            print("Referee", referee, "does not exist.")


def read_sheet(w):
    c = 1
    referrer = w.active.cell(row=1, column=c).value
    while referrer:
        process_column(w, c)
        c += 1
        referrer = w.active.cell(row=1, column=c).value


fn = os.path.join(os.path.join(os.path.expanduser("~")), "commonology/scripts/hacks/", "Kingdoms.xlsx")
ws = openpyxl.load_workbook(fn)

for sheet_name in ws.sheetnames:
    ws.active = ws.sheetnames.index(sheet_name)
    read_sheet(ws)

# Move all from ms@quizitive.com to ms@koplon.com
k = Player.objects.get(email="ms@koplon.com")
q = Player.objects.get(email="ms@quizitive.com")
for p in Player.objects.all():
    if p.referrer == q:
        p.referrer = k
        p.save()
