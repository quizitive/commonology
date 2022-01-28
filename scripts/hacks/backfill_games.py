
import sys
import os
import django
from django.db import transaction
from dateutil.relativedelta import relativedelta
import openpyxl


path = os.getcwd()
sys.path.append(path)
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "project.settings")
django.setup()

from django.db.models import Min
from collections import defaultdict
from project.utils import our_now
from users.models import Player
from users.utils import player_log_entry
from game.models import Game, Series, Answer
from leaderboard.models import PlayerRankScore

fn = os.path.expanduser('~/Downloads')
fn = os.path.join(os.path.join(fn, 'CommonologyBackfillWeeks.xlsx'))
ws = openpyxl.load_workbook(fn)

commonology = Series.objects.get(slug="commonology")

game_44 = Game.objects.get(game_id=44)


def calc_dates(game_id):
    start_date = game_44.start
    end_date = game_44.end
    n = (44 - game_id) * 7
    delta = relativedelta(days=n)
    start_date -= delta
    end_date -= delta
    return start_date, end_date


def list_game_dates():
    print("Listing calculated game begin and end dates.")
    for i in range(0, 44):
        b, e = calc_dates(i)
        delta = relativedelta(days=2)
        e += delta
        print(f"{i} {b.date()} -> {e.strftime('%A')} {e.date()}")
    print()
    print()


def set_game_dates(g):
    start_date, end_date = calc_dates(g.game_id)
    g.start = start_date
    g.end = end_date
    return start_date, end_date


def get_player(email, display_name, join_date):
    try:
        p = Player.objects.get(email=email)
    except Player.DoesNotExist:
        p = Player(email=email)
        p.display_name = display_name
        p.subscribed = False
        p.date_joined = join_date
        p.save()
        p.series.add(commonology)
        player_log_entry(p, "Added by backfill_games.py script.")
        print(f"Adding player {display_name} {email}")

    return p


@transaction.atomic
def process_game(game_id, data):
    g = Game(game_id=game_id)
    g.name = f'Commonology Game {game_id}'
    g.series = commonology
    set_game_dates(g)
    g.save()

    lb = g.leaderboard
    lb.publish_date = g.end + relativedelta(days=3)
    lb.save()

    for email, name, score, rank in data:
        p = get_player(email=email, display_name=name, join_date=g.start)
        prs = PlayerRankScore()
        prs.player = p
        prs.leaderboard = lb
        prs.rank = rank
        prs.score = score
        prs.save()


def print_player_scores(game_id):
    print(f"Listing scores and ranks for all players of game {game_id}")
    g = Game.objects.get(game_id=game_id, series=commonology)
    lb = g.leaderboard
    for i in PlayerRankScore.objects.filter(leaderboard=lb).all():
        print(i.player, i.score, i.rank)


def read_sheet(w):
    data = []
    r = 1

    email = w.active.cell(row=r, column=1).value
    if '@' not in email:
        r += 1
        email = w.active.cell(row=r, column=1).value
    while email:
        name = w.active.cell(row=r, column=2).value
        score = w.active.cell(row=r, column=3).value
        data.append([email, name, score, 0])

        r += 1
        email = w.active.cell(row=r, column=1).value

    rank = 0
    last_score = '0'

    missing_scores = [x for x in data if not x[2]]
    if missing_scores:
        print(f'Missing scores for: {missing_scores}')

    data.sort(key=lambda x: -x[2])
    for rec in data:
        email, name, score, x = rec
        if score == last_score:
            rank = rank
        else:
            rank += 1
            last_score = score
        rec[2] = float(score)
        rec[3] = rank

    return data


def process(game_id):

    if 0 == game_id:
        sheet_name = 'Week A1'
    else:
        sheet_name = f'Week {game_id}'

    print(f"Processing {sheet_name}")
    ws.active = ws.sheetnames.index(sheet_name)
    data = read_sheet(ws)

    process_game(game_id, data)


def do_it():
    print('Starting to backfill game results into PlayerRankScore model.')
    for game_id in range(0, 28):
        if Game.objects.filter(series=commonology, game_id=game_id).exists():
            print(f'Skipping game_id {game_id} because it exists.')
        else:
            print(f'Processing game_id {game_id}')
            process(game_id)
            # print_player_scores(game_id)
        print()
        print()


def set_join_dates():
    print('Setting players join dates to first answer record date if their join date is later than that.')
    # Set the join date to the first game played start date for anyone who doesn't have a join date set.
    players = Player.objects.filter(series=commonology).all()
    old_dates = defaultdict(int)
    for p in players:
        first_play_date = Answer.objects.filter(player_id=p.id).aggregate(joined=Min('timestamp'))['joined']
        if first_play_date is None:
            continue

        if p.date_joined > first_play_date:
            old_date = first_play_date.date()
            old_dates[old_date] += 1
            p.date_joined = first_play_date
            p.save()
            msg = "Set date joined to start date of first game played"
            print(f'{msg} for {p.email} it was {old_date}.')
            player_log_entry(p, f"{msg}.")

    print()
    print('Old dates:')
    for d in old_dates.keys():
        print('  ', d)
    print()
    print()


def game_winner_sanity_check():
    print("Looking for any two games that have the same winning score.")
    x = {i.leaderboard.game.game_id: i.score for i in PlayerRankScore.objects.filter(rank=1).all()}
    x = [[k, x[k]] for k in x.keys()]
    x.sort(key=lambda x: -x[1])
    y = [[one[0], one[1], two[0], two[1]] for one, two in zip(x[1:], x[:-1]) if one[1] == two[1]]
    if y:
        for i in y:
            print(f"These games have the same winner score: {i[0]} and {i[2]}")
    else:
        print('None of the games have the same winning score.')
    print()
    print()


def list_winners():
    print("Listing game winners for all games.")
    winners = [[i.leaderboard.game.game_id, str(i.leaderboard.game.start.date()),
                i.player.email, i.score, i.rank] for i in
               PlayerRankScore.objects.filter(rank=1, leaderboard__game__series=commonology).all()]
    winners.sort()
    for w in winners:
        print(w)
    print()
    print()


do_it()
list_game_dates()
set_join_dates()
game_winner_sanity_check()
list_winners()
# print_player_scores(0)

#  pg_restore -d postgresql://postgres:postgres@localhost/commonology --verbose --clean --no-acl --no-owner ./pg_dumps/commonology_Tue.tar
